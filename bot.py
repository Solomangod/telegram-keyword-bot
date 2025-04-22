import os
from telegram import Update, InputFile
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from openpyxl import load_workbook
from io import BytesIO
from dotenv import load_dotenv

print("🚀 Khởi động bot...")

load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
print("🔑 TOKEN:", TOKEN)

if not TOKEN:
    print("❌ Không tìm thấy BOT_TOKEN trong .env")
    exit()

user_data = {}
stop_flags = {}

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        """Gửi file .txt chứa từ khóa trước, sau đó gửi file Excel .xlsx.
Gõ /stop để dừng quá trình xử lý."""
    )

async def stop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    stop_flags[chat_id] = True
    await update.message.reply_text("⏹ Đã gửi yêu cầu dừng.")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    file_name = doc.file_name.lower()

    if file_name.endswith(".txt"):
        await handle_txt(update, context)
    elif file_name.endswith(".xlsx"):
        await handle_excel(update, context)
    else:
        await update.message.reply_text("Chỉ hỗ trợ file .txt và .xlsx")

async def handle_txt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    file = await update.message.document.get_file()
    content = await file.download_as_bytearray()
    keywords = set(kw.strip().lower() for kw in content.decode("utf-8").splitlines() if kw.strip())
    user_data[update.message.chat_id] = {"keywords": keywords}
    await update.message.reply_text(f"✅ Đã nhận {len(keywords)} từ khóa. Gửi file Excel tiếp theo.")

async def handle_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.message.chat_id
    stop_flags[chat_id] = False

    if chat_id not in user_data:
        await update.message.reply_text("⚠️ Bạn chưa gửi file .txt chứa từ khóa.")
        return

    file = await update.message.document.get_file()
    xlsx_bytes = await file.download_as_bytearray()

    wb = load_workbook(filename=BytesIO(xlsx_bytes))
    ws = wb.active
    keywords = user_data[chat_id]["keywords"]

    progress_message = await update.message.reply_text("🔄 Đang xử lý file...")

    
    total = sum(1 for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value)
    
    match_count = 0

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1), start=1):
        if stop_flags.get(chat_id):
            await update.message.reply_text("⏹ Đã dừng theo yêu cầu.")
            return
        cell = row[0].value
        text = str(cell).lower() if cell else ""
        words = set(text.replace(",", " ").replace(".", " ").replace("!", " ").replace("?", " ").split())
        found = any(kw in words for kw in keywords)
        row[0].offset(column=6).value = "SOS: Nó kia kìa nó kia kìa ÐTH" if found else ""
        if found:
            match_count += 1

        if idx % max(1, total // 20) == 0 or idx == total:
            percent = int((idx / total) * 100)
            await progress_message.edit_text(f"🔄 Đang xử lý: {percent}%")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    await progress_message.edit_text(f"✅ Hoàn tất: {total} dòng, {match_count} dòng gắn SOS.")
    await update.message.reply_document(document=InputFile(output, filename="Checked_Results.xlsx"))

app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("stop", stop))
app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

print("✅ Bot sẵn sàng!")
app.run_polling()